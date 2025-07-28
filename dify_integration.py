"""
Customs Declaration Plugin for Dify
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

这是一个用于处理物流报关文件的Dify插件，可自动处理装箱单、发票和随附文件，
并生成标准化的报关Excel文件。

:copyright: (c) 2025 迈创企业管理服务股份有限公司
:license: MIT, see LICENSE for more details.
"""
import os
import re
import uuid
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font

# 配置部分 - 在Dify环境中设置
UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"
TEMPLATES_FOLDER = "templates"
IMAGES_FOLDER = "images"

# 确保目录存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# 示例模板文件路径 (在实际部署时需要提供)
EXAMPLE_FILE = os.path.join(TEMPLATES_FOLDER, "001.xlsx")

def process_packing_list_files(upload_folder):
    """处理Packing List文件，第一个文件特殊处理，其余文件追加。"""
    all_dfs_to_concat = []
    is_first_file = True

    packing_list_files = sorted(
        [
            f
            for f in os.listdir(upload_folder)
            if re.sub(r"\s+", " ", f).endswith(" Packing List.xlsx")
        ]
    )

    for file in packing_list_files:
        file_path = os.path.join(upload_folder, file)

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        if not ws:
            continue
        bol = ws["K5"].value if ws.cell(row=5, column=11).value else ""

        df_from_row_12 = pd.read_excel(
            file_path, skiprows=11, usecols="B:P", header=None, dtype=object
        )

        total_row = df_from_row_12[
            df_from_row_12.iloc[:, 0]
            .astype(str)
            .str.contains("TOTAL", na=False, case=False)
        ].index
        if not total_row.empty:
            df_from_row_12 = df_from_row_12.iloc[: total_row[0]]

        df_from_row_12.dropna(how="all", inplace=True)

        if is_first_file:
            row_11_data = [
                ws.cell(row=11, column=col).value for col in range(2, 17)
            ]  # B to P
            df_row_11 = pd.DataFrame([row_11_data], columns=df_from_row_12.columns)
            current_file_df = pd.concat([df_row_11, df_from_row_12], ignore_index=True)
            is_first_file = False
        else:
            current_file_df = df_from_row_12

        if not current_file_df.empty:
            current_file_df.columns = range(15)
            current_file_df["BOL"] = bol
            current_file_df["收货地址"] = ""
            all_dfs_to_concat.append(current_file_df)

    if all_dfs_to_concat:
        final_df = pd.concat(all_dfs_to_concat, ignore_index=True)
        return final_df
    else:
        return pd.DataFrame(
            columns=pd.Index([str(i) for i in range(15)] + ["BOL", "收货地址"])
        )


def process_invoice_files(upload_folder):
    """处理Invoice文件"""
    inv_columns = [
        "P/N", "DESCRIPTION", "HS", "NAME", "UNIT", "Q'TY (SET)",
        "U/P (USD)", "AMOUNT (USD)", "NW", "GW", "IsKits",
    ]
    all_data = []

    for file in os.listdir(upload_folder):
        if re.sub(r"\s+", " ", file).endswith(" HIC Invoice.xlsx"):
            file_path = os.path.join(upload_folder, file)
            wb_inv = openpyxl.load_workbook(file_path)
            ws = wb_inv.active
            if not ws:
                continue
            invoice_no_val = ws["G1"].value
            invoice_no = (
                invoice_no_val.replace("INVOICE NO.", "") if invoice_no_val else ""
            )
            df = pd.read_excel(file_path, skiprows=12, usecols="B:L")
            total_row = df[
                df.iloc[:, 4].astype(str).str.contains("Total", na=False, case=False)
            ].index
            if not total_row.empty:
                df = df.iloc[: total_row[0]]
            df.columns = inv_columns
            df["BOL"] = invoice_no
            all_data.append(df)

    return (
        pd.concat(all_data, ignore_index=True)
        if all_data
        else pd.DataFrame(columns=pd.Index(inv_columns + ["BOL"]))
    )


def process_contract_data(inv_data):
    """生成合同表数据"""
    contract_columns = ["品名", "料号", "单位", "数量", "单价(USD)", "总额(USD)"]
    selected_columns = ["P/N", "NAME", "Q'TY (SET)", "U/P (USD)", "AMOUNT (USD)"]
    df = inv_data[selected_columns].copy()
    df.columns = ["品名", "料号", "数量", "单价(USD)", "总额(USD)"]
    df.insert(2, "单位", "")
    return df[contract_columns]


def process_declaration_files(upload_folder):
    """处理随附文件"""
    decl_columns = [
        "Item", "Ordered Qty", "中文品名", "HS", "mag", "是否含电池",
        "鉴定证书编号", "证书类型", "DG", "BOL", "申报要素",
    ]
    all_data = []

    for file in os.listdir(upload_folder):
        if file.endswith("随附文件.xlsx"):
            file_path = os.path.join(upload_folder, file)
            df = pd.read_excel(file_path, skiprows=0)
            df = df[decl_columns]
            all_data.append(df)

    return (
        pd.concat(all_data, ignore_index=True)
        if all_data
        else pd.DataFrame(columns=pd.Index(decl_columns))
    )


def merge_bol_cells(ws, data, bol_col, start_row):
    """合并相同BOL值的单元格"""
    current_bol = None
    start_merge_row = start_row
    for idx, row in data.iterrows():
        bol = row["BOL"]
        if current_bol is None:
            current_bol = bol
            start_merge_row = idx + start_row
        elif bol != current_bol:
            if idx + start_row - 1 > start_merge_row:
                ws.merge_cells(
                    f"{bol_col}{start_merge_row}:{bol_col}{idx + start_row - 1}"
                )
            current_bol = bol
            start_merge_row = idx + start_row
    if current_bol is not None and len(data) > 0 and idx + start_row > start_merge_row:
        ws.merge_cells(f"{bol_col}{start_merge_row}:{bol_col}{idx + start_row}")


def generate_output_excel(
    pkl_data, inv_data, contract_data, decl_data, output_path, id_str, date_str
):
    """生成输出Excel文件"""
    if not os.path.exists(EXAMPLE_FILE):
        raise FileNotFoundError(f"Template file not found at: {EXAMPLE_FILE}")

    try:
        wb = openpyxl.load_workbook(EXAMPLE_FILE)
    except OSError as e:
        raise OSError(f"Failed to load template file at {EXAMPLE_FILE}: {e}")

    # PKL表
    try:
        ws_pkl = wb["PKL "]
        ws_pkl["J10"] = id_str
        ws_pkl["J11"] = date_str
    except KeyError:
        raise KeyError(
            f"Worksheet 'PKL ' not found in {EXAMPLE_FILE}. Available sheets: {wb.sheetnames}"
        )

    image_01_path = os.path.join(IMAGES_FOLDER, "01.png")
    if os.path.exists(image_01_path):
        img1 = Image(image_01_path)
        ws_pkl.add_image(img1, "A1")

    if not pkl_data.empty:
        start_row_pkl = 14
        end_row_pkl = start_row_pkl + len(pkl_data)
        for merged_range in list(ws_pkl.merged_cells.ranges):
            if (
                merged_range.max_row >= start_row_pkl
                and merged_range.min_row <= end_row_pkl
            ):
                ws_pkl.unmerge_cells(str(merged_range))

        header_row = pkl_data.iloc[0]
        data_rows = pkl_data.iloc[1:].reset_index(drop=True)

        for col_idx in range(15):
            if col_idx < len(header_row.iloc[:-2]):
                ws_pkl.cell(row=14, column=col_idx + 1).value = header_row.iloc[col_idx]
        ws_pkl.cell(row=14, column=16).value = "BOL"
        ws_pkl.cell(row=14, column=17).value = "收货地址"

        for row_idx, row_data in data_rows.iterrows():
            for col_idx in range(15):
                if col_idx < len(row_data.iloc[:-2]):
                    ws_pkl.cell(
                        row=row_idx + 15, column=col_idx + 1
                    ).value = row_data.iloc[col_idx]
            ws_pkl.cell(row=row_idx + 15, column=16).value = row_data["BOL"]
            ws_pkl.cell(row=row_idx + 15, column=17).value = row_data["收货地址"]
        
        merge_bol_cells(ws_pkl, data_rows, "P", 15)
    else:
        ws_pkl.cell(row=14, column=16).value = "BOL"
        ws_pkl.cell(row=14, column=17).value = "收货地址"

    # INV表
    try:
        ws_inv = wb["INV"]
        ws_inv["G14"] = id_str
        ws_inv["G15"] = date_str
    except KeyError:
        raise KeyError(
            f"Worksheet 'INV' not found in {EXAMPLE_FILE}. Available sheets: {wb.sheetnames}"
        )

    image_02_path = os.path.join(IMAGES_FOLDER, "02.png")
    if os.path.exists(image_02_path):
        img2 = Image(image_02_path)
        ws_inv.add_image(img2, "A1")

    if not inv_data.empty:
        start_row_inv = 19
        end_row_inv = start_row_inv + len(inv_data)
        for merged_range in list(ws_inv.merged_cells.ranges):
            if (
                merged_range.min_row >= start_row_inv
                and merged_range.min_row <= end_row_inv
            ):
                ws_inv.unmerge_cells(str(merged_range))

    inv_columns = [
        "P/N", "DESCRIPTION", "HS", "NAME", "UNIT", "Q'TY (SET)", "U/P (USD)",
        "AMOUNT (USD)", "NW", "GW", "IsKits", "BOL",
    ]
    for col, header in enumerate(inv_columns, 1):
        ws_inv.cell(row=18, column=col).value = header
    for row_idx, row in inv_data.iterrows():
        for col_idx, value in enumerate(row, 1):
            ws_inv.cell(row=row_idx + 19, column=col_idx).value = value
    merge_bol_cells(ws_inv, inv_data, "L", 19)

    # 合同表
    try:
        ws_contract = wb["合同 "]
        ws_contract["F2"] = id_str
        ws_contract["F3"] = date_str
    except KeyError:
        raise KeyError(
            f"Worksheet '合同 ' not found in {EXAMPLE_FILE}. Available sheets: {wb.sheetnames}"
        )
    contract_columns = ["品名", "料号", "单位", "数量", "单价(USD)", "总额(USD)"]
    for col, header in enumerate(contract_columns, 1):
        ws_contract.cell(row=11, column=col).value = header
    for row_idx, row in contract_data.iterrows():
        for col_idx, value in enumerate(row, 1):
            ws_contract.cell(row=row_idx + 12, column=col_idx).value = value

    # 申报要素表
    try:
        ws_decl = wb["申报要素"]
    except KeyError:
        raise KeyError(
            f"Worksheet '申报要素' not found in {EXAMPLE_FILE}. Available sheets: {wb.sheetnames}"
        )
    decl_columns = [
        "Item", "Ordered Qty", "中文品名", "HS", "mag", "是否含电池",
        "鉴定证书编号", "证书类型", "DG", "BOL", "申报要素",
    ]
    for col, header in enumerate(decl_columns, 1):
        ws_decl.cell(row=1, column=col).value = header
    for row_idx, row in decl_data.iterrows():
        for col_idx, value in enumerate(row, 1):
            ws_decl.cell(row=row_idx + 2, column=col_idx).value = value

    # 将所有单元格内容居中
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_path)
    return output_path


def process_customs_declaration(input_folder_path, output_folder_path=None):
    """
    Dify集成主函数
    
    Args:
        input_folder_path (str): 包含待处理Excel文件的文件夹路径
        output_folder_path (str, optional): 输出文件夹路径，默认为None时使用默认路径
        
    Returns:
        dict: 处理结果信息
    """
    # 如果没有指定输出文件夹，则使用默认路径
    if output_folder_path is None:
        output_folder_path = OUTPUT_FOLDER
    
    # 确保输出文件夹存在
    os.makedirs(output_folder_path, exist_ok=True)
    
    # 处理各种文件
    pkl_data = process_packing_list_files(input_folder_path)
    inv_data = process_invoice_files(input_folder_path)
    contract_data = process_contract_data(inv_data)
    decl_data = process_declaration_files(input_folder_path)
    
    # 生成输出文件名和ID
    today = datetime.now()
    date_str = (
        today.strftime("%Y/%#m/%#d")
        if os.name == "nt"
        else today.strftime("%Y/%-m/%-d")
    )
    id_str = today.strftime("WWSH%Y%m%d001")
    output_filename = f"{id_str}-苏州吴江提货.xlsx"
    output_path = os.path.join(output_folder_path, output_filename)
    
    # 生成输出Excel文件
    try:
        generate_output_excel(
            pkl_data, inv_data, contract_data, decl_data, output_path, id_str, date_str
        )
        
        return {
            "status": "success",
            "message": "报关文件处理完成",
            "output_file": output_path,
            "file_id": id_str,
            "processing_date": date_str,
            "summary": {
                "packing_list_items": len(pkl_data),
                "invoice_items": len(inv_data),
                "contract_items": len(contract_data),
                "declaration_items": len(decl_data)
            }
        }
    except Exception as e:
        return {
            "status": "error",
            "message": f"处理过程中出现错误: {str(e)}",
            "output_file": None
        }


# Dify代码执行节点可以直接调用的函数示例
def dify_main(input_folder_path):
    """
    Dify平台入口函数
    
    Args:
        input_folder_path (str): 输入文件夹路径
        
    Returns:
        dict: 处理结果
    """
    return process_customs_declaration(input_folder_path)
